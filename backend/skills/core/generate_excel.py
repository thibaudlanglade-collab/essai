"""
Skill: generate_excel
Purpose: Generate an .xlsx file from structured data. No LLM.
         Accepts single-sheet (list of dicts) or multi-sheet (list of tables).
"""
from __future__ import annotations

import io
from typing import Any

from skills.base import SkillResult

SKILL_ID = "generate_excel"
DESCRIPTION = "Generate an .xlsx file from structured data (single sheet or multiple sheets)"

TOOL_SCHEMA = {
    "name": "generate_excel",
    "description": (
        "Generate an Excel file (.xlsx bytes) from structured data. "
        "Accepts either a single list of dicts (one sheet) or a list of table "
        "dicts with 'name', 'columns', 'rows' (multi-sheet). Returns the xlsx "
        "bytes ready to be served as a download."
    ),
    "when_to_use": [
        "Final step to produce a downloadable Excel file",
        "User wants the result as a spreadsheet",
    ],
    "when_not_to_use": [
        "User wants CSV, JSON, or another format",
        "Data is not yet structured",
    ],
    "input_schema": {
        "type": "object",
        "properties": {
            "data_ref": {
                "type": "string",
                "description": "State key reference to the data (list of dicts OR list of tables)",
            },
            "default_sheet_name": {
                "type": "string",
                "description": "Sheet name when input is a single list of dicts (default: 'Sheet1')",
            },
        },
        "required": ["data_ref"],
    },
}

# Excel sheet name constraints
_FORBIDDEN_CHARS = str.maketrans({c: "_" for c in r":\/?*[]"})
_MAX_SHEET_NAME_LEN = 31


def _sanitize_sheet_name(raw: str, used_names: set[str]) -> str:
    name = raw.translate(_FORBIDDEN_CHARS).strip()
    name = name[:_MAX_SHEET_NAME_LEN]
    if not name:
        name = "sheet"

    # Deduplicate case-insensitively
    lower_used = {n.lower() for n in used_names}
    if name.lower() not in lower_used:
        used_names.add(name)
        return name

    for suffix in range(2, 1000):
        candidate = f"{name[:_MAX_SHEET_NAME_LEN - len(str(suffix)) - 1]}_{suffix}"
        if candidate.lower() not in lower_used:
            used_names.add(candidate)
            return candidate

    # Extremely unlikely fallback
    used_names.add(name)
    return name


def _is_multi_sheet_format(data: list) -> bool:
    """True if the list looks like a list of table objects (name/columns/rows)."""
    if not data:
        return False
    first = data[0]
    return isinstance(first, dict) and "columns" in first and "rows" in first


async def execute(input_data: dict, context: Any) -> SkillResult:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return SkillResult(
            success=False,
            data=None,
            error="generate_excel: openpyxl is not installed",
        )

    try:
        data: Any = input_data.get("data_ref")
        default_sheet_name: str = input_data.get("default_sheet_name", "Sheet1")

        if not isinstance(data, list):
            return SkillResult(
                success=False,
                data=None,
                error=(
                    f"generate_excel: 'data_ref' must resolve to a list, "
                    f"got {type(data).__name__}"
                ),
            )

        if len(data) == 0:
            return SkillResult(
                success=False,
                data=None,
                error="generate_excel: data list is empty — nothing to write",
            )

        wb = openpyxl.Workbook()
        # Remove the default empty sheet created by openpyxl
        if wb.active is not None:
            wb.remove(wb.active)

        used_names: set[str] = set()
        bold_arial = Font(name="Arial", bold=True)
        normal_arial = Font(name="Arial")

        # ── Multi-sheet mode ──────────────────────────────────────────────────
        if _is_multi_sheet_format(data):
            total_rows = 0
            for table in data:
                if not isinstance(table, dict):
                    continue
                raw_name = str(table.get("name", "table"))
                columns: list[str] = table.get("columns", [])
                rows: list[dict] = table.get("rows", [])

                sheet_name = _sanitize_sheet_name(raw_name, used_names)
                ws = wb.create_sheet(title=sheet_name)

                # Header row
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = bold_arial

                # Data rows
                for row_idx, row_dict in enumerate(rows, start=2):
                    for col_idx, col_name in enumerate(columns, start=1):
                        cell = ws.cell(
                            row=row_idx,
                            column=col_idx,
                            value=row_dict.get(col_name),
                        )
                        cell.font = normal_arial

                total_rows += len(rows)

            mode = "multi"
            sheets_written = len(wb.sheetnames)

        # ── Single-sheet mode ─────────────────────────────────────────────────
        else:
            if not isinstance(data[0], dict):
                return SkillResult(
                    success=False,
                    data=None,
                    error=(
                        "generate_excel: expected list of dicts or list of table objects, "
                        f"got list of {type(data[0]).__name__}"
                    ),
                )

            sheet_name = _sanitize_sheet_name(default_sheet_name, used_names)
            ws = wb.create_sheet(title=sheet_name)
            columns = list(data[0].keys())

            # Header row
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = bold_arial

            # Data rows
            for row_idx, row_dict in enumerate(data, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=row_dict.get(col_name),
                    )
                    cell.font = normal_arial

            mode = "single"
            sheets_written = 1
            total_rows = len(data)

        # ── Serialise ─────────────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()

        return SkillResult(
            success=True,
            data=xlsx_bytes,
            debug={
                "mode": mode,
                "sheets_written": sheets_written,
                "total_rows": total_rows,
            },
        )

    except Exception as exc:
        return SkillResult(
            success=False,
            data=None,
            error=f"generate_excel failed: {exc}",
        )

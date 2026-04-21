"""Append an invoice row to the tenant's yearly `Suivi_Factures_<YYYY>.xlsx`.

Kept as a tiny service (not a skill) because it takes a fully-resolved
filesystem path that a planner would never be allowed to compose on its
own. The file_organizer orchestrator computes the path per tenant and
calls this function.

Dedup is based on `(invoice_number, supplier_name)` — re-saving the same
facture (e.g. when the prospect edits + re-saves) does not spam the file.
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional


logger = logging.getLogger(__name__)


_EXCEL_COLUMNS = [
    "Date",
    "Fournisseur",
    "N° facture",
    "Montant HT",
    "TVA",
    "Montant TTC",
    "Auto-liquidation",
    "Dossier",
    "Fichier",
    "Importé le",
]


def append_invoice_row(
    *,
    excel_path: Path,
    invoice_data: dict[str, Any],
    target_folder: Optional[str],
    final_filename: str,
) -> tuple[bool, int]:
    """Append one invoice row. Returns (added, total_rows).

    `added=False` means the row was a dedup hit against an existing
    `(invoice_number, supplier_name)`; no write happened.
    """
    from openpyxl import Workbook, load_workbook

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb.active
        key = (
            (invoice_data.get("invoice_number") or "").strip().lower(),
            (invoice_data.get("supplier_name") or "").strip().lower(),
        )
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_key = (
                str(row[2] or "").strip().lower(),
                str(row[1] or "").strip().lower(),
            )
            if row_key == key and any(key):
                return False, max(0, ws.max_row - 1)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Factures"
        ws.append(_EXCEL_COLUMNS)

    ws.append(
        [
            _format_date(invoice_data.get("invoice_date")),
            invoice_data.get("supplier_name") or "",
            invoice_data.get("invoice_number") or "",
            _format_number(invoice_data.get("amount_ht")),
            _format_number(invoice_data.get("amount_vat")),
            _format_number(invoice_data.get("amount_ttc")),
            "Oui" if invoice_data.get("auto_liquidation") else "Non",
            (target_folder or "").rstrip("/"),
            final_filename or "",
            datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        ]
    )
    wb.save(excel_path)
    return True, max(0, ws.max_row - 1)


def invoice_year(invoice_data: dict[str, Any]) -> int:
    raw = invoice_data.get("invoice_date")
    if isinstance(raw, date):
        return raw.year
    if isinstance(raw, str) and len(raw) >= 4:
        try:
            return int(raw[:4])
        except ValueError:
            pass
    return datetime.utcnow().year


def _format_date(value: Any) -> str:
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value[:10]
    return ""


def _format_number(value: Any) -> Any:
    if value is None or value == "":
        return ""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return str(value)

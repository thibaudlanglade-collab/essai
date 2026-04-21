"""Per-tenant file classement + Excel tracking (brief §6.1, volet 1).

When a prospect validates an extraction in Smart Extract and flips
`commit_to_invoices`, the stored file needs to be:
1. Moved from `attachments/{user_id}/extractions/` to
   `attachments/{user_id}/organized/{target_folder}/{final_filename}`
   so the tenant's filesystem reflects the suggested classement.
2. Reflected as a row in the yearly Excel `Suivi_Factures_{YYYY}.xlsx`
   kept under `attachments/{user_id}/`.

Returns a list of ordered "steps" — short human sentences — that the
frontend streams as a live toast ("Fichier renommé", "Classé dans
Factures/Point_P/Avril_2026/", "Excel mis à jour"). Makes the automation
visible rather than invisible.

The operations are all filesystem-local to the prospect's sandbox. Drive
integration (volet 2) is layered on top via `services/drive_service.py`
— this file never reaches outside `attachments/{user_id}/`.
"""
from __future__ import annotations

import logging
import shutil
import unicodedata
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional


logger = logging.getLogger(__name__)


_ATTACHMENTS_ROOT = Path(__file__).resolve().parent.parent / "attachments"


_EXCEL_COLUMNS = [
    "Date",
    "Fournisseur",
    "N° facture",
    "Montant HT",
    "TVA",
    "Montant TTC",
    "Auto-liquidation",
    "Dossier",
    "Fichier",
    "Importé le",
]


# ─────────────────────────────────────────────────────────────────────────────
# Path helpers
# ─────────────────────────────────────────────────────────────────────────────


def _sanitise_path_segment(value: str) -> str:
    """Make a folder/filename component safe for every filesystem we care about.

    Strips accents, collapses whitespace/separators to `_`, drops characters
    that are unsafe on Windows (the prospects will sync this back to their
    own machines eventually).
    """
    nfd = unicodedata.normalize("NFD", value)
    ascii_only = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", ascii_only)
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = cleaned.strip("._")
    return cleaned or "document"


def _sanitise_relative_folder(folder: str) -> Path:
    """Turn `"Factures/Point_P/Avril_2026/"` into a safe relative `Path`.

    Rejects absolute paths + `..` traversal. The result is always relative
    and stays inside `organized/`.
    """
    raw = (folder or "").strip().strip("/").strip("\\")
    if not raw:
        return Path("Autres")
    parts: list[str] = []
    for segment in raw.replace("\\", "/").split("/"):
        if not segment or segment in {".", ".."}:
            continue
        parts.append(_sanitise_path_segment(segment))
    return Path(*parts) if parts else Path("Autres")


def _sanitise_filename(filename: str, fallback_ext: Optional[str] = None) -> str:
    """Clean a filename; preserve the extension; ensure a `.ext` exists."""
    base = Path(filename or "").name or "document"
    stem = Path(base).stem or "document"
    ext = Path(base).suffix
    if not ext:
        ext = fallback_ext or ".pdf"
    cleaned_stem = _sanitise_path_segment(stem)
    if not ext.startswith("."):
        ext = "." + ext
    return f"{cleaned_stem}{ext}"


def _unique_destination(dest: Path) -> Path:
    """If `dest` already exists, append ` (1)`, ` (2)`… before the extension."""
    if not dest.exists():
        return dest
    stem = dest.stem
    ext = dest.suffix
    parent = dest.parent
    for n in range(1, 100):
        candidate = parent / f"{stem} ({n}){ext}"
        if not candidate.exists():
            return candidate
    raise OSError(f"Cannot find a unique filename for {dest} after 100 tries.")


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────


def organize_invoice_file(
    *,
    user_id: str,
    source_path: Optional[Path | str],
    target_folder: str,
    final_filename: str,
    invoice_data: dict[str, Any],
) -> dict[str, Any]:
    """Move the stored file into its target folder + update the yearly Excel.

    Best-effort: if the source file is missing (e.g. text-only extraction),
    we still update the Excel and skip the move. Exceptions are caught and
    converted into a failure step so the caller can surface them to the
    prospect without crashing the save.

    Returns:
        {
          "steps": [str, ...],      # ordered human-readable progress
          "final_path": str | None, # absolute path of the moved file
          "excel_path": str | None, # absolute path of the suivi file
          "ok": bool,
        }
    """
    steps: list[str] = []
    user_root = _ATTACHMENTS_ROOT / user_id
    user_root.mkdir(parents=True, exist_ok=True)

    final_path: Optional[Path] = None
    excel_path: Optional[Path] = None

    # ── 1. Move the file ───────────────────────────────────────────────────
    if source_path:
        try:
            src = Path(source_path)
            if not src.is_absolute():
                src = user_root / "extractions" / src.name
            if not src.exists():
                steps.append(
                    f"Fichier introuvable à l'emplacement initial ({src.name}), "
                    "étape de déplacement ignorée."
                )
            else:
                rel_folder = _sanitise_relative_folder(target_folder)
                dest_dir = user_root / "organized" / rel_folder
                dest_dir.mkdir(parents=True, exist_ok=True)
                dest_name = _sanitise_filename(final_filename, fallback_ext=src.suffix)
                dest = _unique_destination(dest_dir / dest_name)
                shutil.move(str(src), str(dest))
                final_path = dest
                steps.append(f"Renommé en {dest.name}.")
                steps.append(f"Classé dans {rel_folder.as_posix()}/.")
        except Exception as exc:
            logger.warning("file move failed for user=%s: %s", user_id, exc, exc_info=True)
            steps.append(f"Déplacement du fichier impossible : {exc}.")
    else:
        steps.append(
            "Aucun fichier joint (document saisi au format texte), "
            "étape de déplacement sautée."
        )

    # ── 2. Update the Excel ────────────────────────────────────────────────
    try:
        year = _invoice_year(invoice_data)
        excel_path = user_root / f"Suivi_Factures_{year}.xlsx"
        added = _append_invoice_row(
            excel_path=excel_path,
            invoice_data=invoice_data,
            target_folder=target_folder,
            final_filename=(final_path.name if final_path else final_filename),
        )
        if added:
            steps.append(
                f"Suivi_Factures_{year}.xlsx mis à jour (+1 ligne, "
                f"{_invoice_count(excel_path)} lignes au total)."
            )
        else:
            steps.append(f"Suivi_Factures_{year}.xlsx déjà à jour (doublon détecté).")
    except Exception as exc:
        logger.warning(
            "excel update failed for user=%s: %s", user_id, exc, exc_info=True
        )
        steps.append(f"Mise à jour de l'Excel impossible : {exc}.")

    return {
        "steps": steps,
        "final_path": str(final_path) if final_path else None,
        "excel_path": str(excel_path) if excel_path else None,
        "ok": bool(steps) and all("impossible" not in s.lower() for s in steps),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────


def _invoice_year(invoice_data: dict[str, Any]) -> int:
    raw = invoice_data.get("invoice_date")
    if isinstance(raw, date):
        return raw.year
    if isinstance(raw, str) and len(raw) >= 4:
        try:
            return int(raw[:4])
        except ValueError:
            pass
    return datetime.utcnow().year


def _append_invoice_row(
    *,
    excel_path: Path,
    invoice_data: dict[str, Any],
    target_folder: str,
    final_filename: str,
) -> bool:
    """Append one invoice row. Returns True when the row was actually added."""
    from openpyxl import Workbook, load_workbook

    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb.active
        # Detect duplicates by (invoice_number, supplier_name).
        key = (
            (invoice_data.get("invoice_number") or "").strip().lower(),
            (invoice_data.get("supplier_name") or "").strip().lower(),
        )
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_key = (
                str(row[2] or "").strip().lower(),
                str(row[1] or "").strip().lower(),
            )
            if row_key == key and any(key):
                return False
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Factures"
        ws.append(_EXCEL_COLUMNS)

    ws.append(
        [
            _format_date(invoice_data.get("invoice_date")),
            invoice_data.get("supplier_name") or "",
            invoice_data.get("invoice_number") or "",
            _format_number(invoice_data.get("amount_ht")),
            _format_number(invoice_data.get("amount_vat")),
            _format_number(invoice_data.get("amount_ttc")),
            "Oui" if invoice_data.get("auto_liquidation") else "Non",
            (target_folder or "").rstrip("/"),
            final_filename or "",
            datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        ]
    )
    wb.save(excel_path)
    return True


def _invoice_count(excel_path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    # -1 for the header row.
    count = max(0, ws.max_row - 1)
    wb.close()
    return count


def _format_date(value: Any) -> str:
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value[:10]
    return ""


def _format_number(value: Any) -> Any:
    if value is None or value == "":
        return ""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return str(value)
